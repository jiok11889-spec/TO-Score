import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# 현행 티어 (3차 대회 5/17 반영 기준, 기존 로직 그대로)
data = [
    # tier, name, is_pro, a 누적평균, b 26년평균, 기준스코어
    (0, '김산',       True,   78.2,  80.7,  79.5),
    (0, '나한영',     False,  84.0,  82.0,  83.0),
    (0, '홍경택',     False,  83.3,  83.0,  83.2),
    (0, '이현우',     True,   None,  None,  None),
    (0, '이형석',     True,   None,  None,  None),
    (0, '임승언',     True,   None,  None,  None),
    (0, '윤석원',     True,   None,  None,  None),
    (1, '이정승',     False,  81.0,  None,  81.0),
    (1, '최창환',     False,  82.0,  82.0,  82.0),
    (1, '김수민',     False,  85.2,  82.0,  83.6),
    (1, '박병우',     False,  84.3,  86.0,  85.2),
    (1, '김대우',     False,  87.0,  None,  87.0),
    (1, '임익준',     False,  90.1,  89.0,  89.5),
    (1, '김성환',     False,  89.3,  90.5,  89.9),
    (2, '이은광',     False,  85.0,  85.0,  85.0),
    (2, '윤준원',     False,  88.0,  88.0,  88.0),
    (2, '강현수',     False,  89.0,  89.0,  89.0),
    (2, '황인호',     False,  87.7,  94.0,  90.8),
    (2, '윤현호',     False,  93.0,  None,  93.0),
    (2, '류성준',     False,  98.7,  90.0,  94.3),
    (2, '김성환(92)', False,  94.5,  95.0,  94.8),
    (2, '조현민',     False,  96.7,  98.5,  97.6),
    (2, '권기표',     False,  97.1,  99.3,  98.2),
    (2, '손원우',     False,  99.0,  None,  99.0),
    (2, '조성태',     False,  None,  None,  None),
    (3, '이준범',     False,  94.0,  91.0,  92.5),
    (3, '이건희',     False,  97.2,  98.5,  97.8),
    (3, '김준기',     False, 100.6,  99.0,  99.8),
    (3, '황호연',     False, 102.7,  98.0, 100.3),
    (3, '백근호',     False, 100.5, 102.0, 101.2),
    (4, '김지수',     False, 102.0,  91.0,  96.5),
    (4, '강충현',     False,  98.5,  98.5,  98.5),
    (4, '이호준',     False, 102.5, 100.0, 101.2),
    (4, '박헌우',     False, 100.0, 104.0, 102.0),
    (4, '김현섭',     False, 102.0, 102.0, 102.0),
    (4, '최원재',     False, 102.5, 103.0, 102.8),
    (4, '김서래',     False, 103.0,  None, 103.0),
    (5, '이도영',     False,  97.7, 103.0, 100.3),
    (5, '옥지엽',     False, 101.3, 110.0, 105.7),
    (5, '서정권',     False, 107.0, 106.0, 106.5),
    (5, '주홍석',     False, 115.0, 115.0, 115.0),
]

tier_counts = {0:7, 1:7, 2:11, 3:5, 4:7, 5:4}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '기존 티어 현황'

tier_header_fill = {
    0: ('1A1A2E', 'FFFFFF'),
    1: ('00695C', 'FFFFFF'),
    2: ('1565C0', 'FFFFFF'),
    3: ('E65100', 'FFFFFF'),
    4: ('6A1B9A', 'FFFFFF'),
    5: ('B71C1C', 'FFFFFF'),
}
tier_row_fill = {
    0: 'ECEFF1',
    1: 'E0F2F1',
    2: 'E3F2FD',
    3: 'FFF3E0',
    4: 'F3E5F5',
    5: 'FFEBEE',
}

thin = Side(style='thin', color='BBBBBB')
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

def cs(cell, value, bold=False, size=10, font_color='000000',
       fill_color=None, align='center', wrap=False):
    cell.value = value
    cell.font = Font(bold=bold, size=size, color=font_color, name='맑은 고딕')
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if fill_color:
        cell.fill = PatternFill('solid', fgColor=fill_color)
    cell.border = border_thin

# ── 제목 ──
ws.merge_cells('A1:F1')
cs(ws['A1'], '사군자 티오방 — 기존 티어 현황 (\'26 3차 대회 5/17 기준)',
   bold=True, size=13, font_color='FFFFFF', fill_color='1A1A2E')
ws.row_dimensions[1].height = 32

ws.merge_cells('A2:F2')
cs(ws['A2'],
   "기준스코어 = 누적평균(a) × 0.5 + 26년평균(b) × 0.5   |   기존 로직 적용 결과   |   T0=7 / T1=7 / T2=11 / T3=5 / T4=7 / T5=4",
   size=9, font_color='444444', fill_color='F0F0F0')
ws.row_dimensions[2].height = 16

# ── 헤더 ──
headers = ['티어', '이름', 'a  누적평균\n(전체 대회)', "b  26년평균\n('26 시즌)", '기준스코어\na×0.5 + b×0.5', '비고']
for col, h in enumerate(headers, 1):
    cs(ws.cell(3, col), h, bold=True, size=10,
       font_color='FFFFFF', fill_color='2C3E50', wrap=True)
ws.row_dimensions[3].height = 38

# ── 데이터 행 ──
for r, (tier, name, is_pro, a, b, ref) in enumerate(data, 4):
    row_bg = tier_row_fill[tier]
    t_bg, t_fg = tier_header_fill[tier]

    tier_label = f'T{tier}' + (' · PRO' if tier == 0 else '')
    cs(ws.cell(r, 1), tier_label, bold=True, font_color=t_fg, fill_color=t_bg)

    name_disp = name + (' ★' if is_pro else '')
    cs(ws.cell(r, 2), name_disp, bold=is_pro, align='left', fill_color=row_bg,
       font_color=('7B6000' if is_pro else '000000'))

    for col, val in [(3, a), (4, b), (5, ref)]:
        disp = f'{val:.1f}' if val is not None else '—'
        fc = '000000'
        if col == 5 and val is not None:
            if val < 87:   fc = '1B5E20'
            elif val < 93: fc = '0D47A1'
            elif val < 99: fc = 'E65100'
            else:          fc = 'B71C1C'
        cs(ws.cell(r, col), disp, fill_color=row_bg, font_color=fc)

    note = 'PRO — 0티어 고정' if is_pro else ('참가 기록 없음' if ref is None else '')
    cs(ws.cell(r, 6), note, size=9, fill_color=row_bg, font_color='999999', align='left')

    ws.row_dimensions[r].height = 20

ws.column_dimensions['A'].width = 11
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 18
ws.freeze_panes = 'A4'

# ── 요약 시트 ──
ws2 = wb.create_sheet('티어별 요약')
ws2.merge_cells('A1:C1')
cs(ws2['A1'], '티어별 인원 요약 (기존 로직)',
   bold=True, size=12, font_color='FFFFFF', fill_color='1A1A2E')
ws2.row_dimensions[1].height = 26

for col, h in enumerate(['티어', '인원', '비고'], 1):
    cs(ws2.cell(2, col), h, bold=True, font_color='FFFFFF', fill_color='2C3E50')

summary = [
    ('T0',   7,  'PRO 5명 + 나한영·홍경택'),
    ('T1',   7,  '이정승·최창환·김수민·박병우·김대우·임익준·김성환'),
    ('T2',  11,  '이은광·윤준원·강현수·황인호·윤현호·류성준·김성환(92)·조현민·권기표·손원우·조성태'),
    ('T3',   5,  '이준범·이건희·김준기·황호연·백근호'),
    ('T4',   7,  '김지수·강충현·이호준·박헌우·김현섭·최원재·김서래'),
    ('T5',   4,  '이도영·옥지엽·서정권·주홍석'),
    ('합계', 41, ''),
]
row_fills2 = ['ECEFF1','E0F2F1','E3F2FD','FFF3E0','F3E5F5','FFEBEE','FAFAFA']
for i, (t, n, rng) in enumerate(summary, 3):
    bg = row_fills2[i - 3]
    for col, val in enumerate([t, n, rng], 1):
        cs(ws2.cell(i, col), val, fill_color=bg,
           align='center' if col < 3 else 'left', bold=(t == '합계'))

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 8
ws2.column_dimensions['C'].width = 60
for r in range(1, 10):
    ws2.row_dimensions[r].height = 20

out_path = 'C:/Users/USER/Documents/TO-Score/기존티어_현황.xlsx'
wb.save(out_path)
print(f'저장 완료: {out_path}')
