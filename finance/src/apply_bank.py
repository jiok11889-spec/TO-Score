"""카카오뱅크 거래내역(암호화 xlsx) → 검증·분류·반영.

    python finance/src/apply_bank.py <거래내역.xlsx> --pw 비밀번호
    python finance/src/apply_bank.py <거래내역.xlsx> --pw 비밀번호 --resolve in/resolve.txt --apply

1회차: 잔액 체인과 장부를 대조하고, 판단이 필요한 건을 **한 번에 모아** 보고한다.
       동시에 `in/resolve.txt` 답안 template을 만든다.
2회차: 그 파일을 채워 --resolve 로 넘기면 입금현황·원장까지 반영한다.

원칙: 추측하지 않는다. 애매하면 멈추고 물어본다. 다만 물어볼 것은 전부 한 번에 묻는다.
"""
import io
import re
import sys
from datetime import date, datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import msoffcrypto
import openpyxl

FIN = Path(__file__).resolve().parent.parent
XLSX = FIN / "data" / "TO Score.xlsx"
RESOLVE_OUT = FIN.parent / "in" / "resolve.txt"
EXPENSE_KINDS = ("모임비", "경조사비", "강연비", "경품비", "지각비")


class Stop(SystemExit):
    pass


# ── 거래내역 읽기 ──────────────────────────────────────────────────────

def read_bank(path, pw):
    buf = io.BytesIO()
    with open(path, "rb") as f:
        of = msoffcrypto.OfficeFile(f)
        of.load_key(password=pw)
        of.decrypt(buf)
    ws = openpyxl.load_workbook(buf, data_only=True).worksheets[0]

    head = None
    txns = []
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        cells = [str(v).strip() if v is not None else "" for v in row]
        if head is None:
            if "거래일시" in cells:
                head = {name: i for i, name in enumerate(cells) if name}
            continue
        dt = cells[head["거래일시"]]
        if not dt:
            continue
        m = re.match(r"(\d{4})\.(\d{2})\.(\d{2})", dt)
        if not m:
            continue
        amt = int(cells[head["거래금액"]].replace(",", ""))
        bal = int(cells[head["거래 후 잔액"]].replace(",", ""))
        txns.append({
            "n": len(txns) + 1,
            "date": date(*map(int, m.groups())),
            "memo": cells[head["내용"]],
            "kind": cells[head["거래구분"]],
            "amount": amt, "balance": bal,
        })
    if not txns:
        raise Stop("거래를 하나도 읽지 못했습니다.")
    return txns


def verify_chain(txns):
    for i in range(1, len(txns)):
        p, c = txns[i - 1], txns[i]
        want = p["balance"] + c["amount"]
        if want != c["balance"]:
            raise Stop(
                f"잔액 체인이 끊깁니다 ({c['n']}번 {c['date']:%m/%d} {c['memo']}).\n"
                f"  {p['balance']:,} + {c['amount']:,} = {want:,} 인데 적힌 잔액은 {c['balance']:,}\n"
                f"  차이 {c['balance']-want:+,} — 거래내역이 누락됐습니다.")
    return txns[0]["balance"] - txns[0]["amount"], txns[-1]["balance"]


# ── 장부 ───────────────────────────────────────────────────────────────

def load_book():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["입금현황"]
    members, retired = {}, set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if not v or str(v).strip() in ("", "계"):
            continue
        name = str(v).strip()
        members[name] = r
        if str(ws.cell(row=r, column=2).value or "").strip() == "탈퇴":
            retired.add(name)
    months = {}
    for c in range(4, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if isinstance(h, str) and "년" in h:
            months[h.strip()] = c
    return wb, ws, members, retired, months


def ym(d):
    return f"{d.year % 100}년 {d.month}월"


def next_ym(label):
    y, m = re.match(r"(\d+)년 (\d+)월", label).groups()
    y, m = int(y), int(m) + 1
    if m > 12:
        y, m = y + 1, 1
    return f"{y}년 {m}월"


def dashboard_balance():
    sys.path.insert(0, str(FIN / "src"))
    import dashboard
    return round(dashboard.build_data()["summary"]["balance"])


def standard_fee(ws, members, months):
    for label in reversed(list(months)):
        vals = [ws.cell(row=r, column=months[label]).value for r in members.values()]
        vals = [v for v in vals if isinstance(v, (int, float)) and v > 0]
        if vals:
            return max(set(vals), key=vals.count)
    return 20000


# ── 분류 ───────────────────────────────────────────────────────────────

def classify(txns, ws, members, retired, months, fee, answers):
    """반영 계획과 '확인 필요' 목록을 함께 만든다. 첫 문제에서 멈추지 않는다."""
    filled = {}          # (name, ym) -> 금액  (기존 + 이번 계획)
    for name, r in members.items():
        for label, c in months.items():
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and v:
                filled[(name, label)] = v

    plan, ledger, asks = [], [], []
    active = [m for m in members if m not in retired]

    def first_free(name, start):
        label = start
        while (name, label) in filled:
            if label not in months:
                raise Stop(f"입금현황에 '{label}' 컬럼이 없습니다 ({name} 선납).")
            label = next_ym(label)
        return label

    def place(name, label, amt):
        if label not in months:
            raise Stop(f"입금현황에 '{label}' 컬럼이 없습니다.")
        filled[(name, label)] = filled.get((name, label), 0) + amt
        plan.append((members[name], months[label], filled[(name, label)], name, label, amt))

    for t in txns:
        a, memo = t["amount"], t["memo"]
        ans = answers.get(t["n"])

        if "이자" in memo or t["kind"] == "예금이자":
            ledger.append((t["date"], "입금", "이자", None, a, memo))
            continue

        if a < 0:
            if ans in EXPENSE_KINDS:
                ledger.append((t["date"], "출금", ans, -a, None, f"{memo}"))
            else:
                asks.append((t, f"지출 {-a:,}원 — 분류를 알려주세요 "
                                f"({'/'.join(EXPENSE_KINDS)})"))
            continue

        cands = [m for m in active if m == memo or m.startswith(memo + "(")]
        if not cands:
            asks.append((t, f"입금 {a:,}원 — 멤버 명단에 '{memo}' 없음 "
                            f"(닉네임·신규·탈퇴자 여부 확인)"))
            continue

        name = None
        if len(cands) == 1:
            name = cands[0]
        elif ans and ans.split()[0] in cands:
            name = ans.split()[0]
        else:
            asks.append((t, f"입금 {a:,}원 — 동명이인 {len(cands)}명 중 누구인지 "
                            f"({', '.join(cands)})"))
            continue

        this = ym(t["date"])
        if a == fee:
            if (name, this) in filled:
                if ans and "중복" in ans:
                    place(name, this, a)
                elif ans and ("선납" in ans or "다음" in ans):
                    place(name, first_free(name, this), a)
                else:
                    asks.append((t, f"{name} {a:,}원 — {this}가 이미 채워져 있음. "
                                    f"중복입금인가요, 다음 달 선납인가요? (중복 / 선납)"))
                continue
            place(name, this, a)
        elif a % fee == 0:
            k = a // fee
            if ans and re.search(r"x\s*(\d+)", ans):
                k = int(re.search(r"x\s*(\d+)", ans).group(1))
            elif not ans:
                asks.append((t, f"{name} {a:,}원 — 회비의 {k}배. "
                                f"{k}개월 선납이 맞나요? (예: '{name} x{k}')"))
                continue
            label = first_free(name, this)
            for _ in range(k):
                place(name, label, fee)
                label = first_free(name, label)
        else:
            if ans and "부분" in ans:
                place(name, this, a)
            else:
                asks.append((t, f"{name} {a:,}원 — 회비({fee:,}) 배수가 아님. "
                                f"부분납부인가요? (부분)"))
    return plan, ledger, asks


# ── 원장 append ────────────────────────────────────────────────────────

def append_ledger(wb, entries):
    import copy
    wl = wb["원장"]
    src = wl.max_row
    for i, (d, k1, k2, cha, dae, memo) in enumerate(entries):
        r = wl.max_row + 1
        for c in range(1, 10):
            wl.cell(row=r, column=c)._style = copy.copy(wl.cell(row=src, column=c)._style)
        wl.cell(row=r, column=1).value = f'=RIGHT(YEAR(C{r}),2)&"년 "&MONTH(C{r})&"월"'
        wl.cell(row=r, column=2).value = f'=VLOOKUP(WEEKDAY(C{r},2),참조!$A$1:$B$10,2,)'
        wl.cell(row=r, column=3).value = datetime(d.year, d.month, d.day)
        wl.cell(row=r, column=4).value = k1
        wl.cell(row=r, column=5).value = k2
        wl.cell(row=r, column=6).value = cha
        wl.cell(row=r, column=7).value = dae
        wl.cell(row=r, column=8).value = f"=G{r}-F{r}"
        wl.cell(row=r, column=9).value = memo


def read_answers(path):
    ans = {}
    if not path:
        return ans
    for line in Path(path).read_text(encoding="utf-8").splitlines():
        line = line.split("#")[0].strip()
        if not line or "=" not in line:
            continue
        k, v = line.split("=", 1)
        if k.strip().isdigit():
            ans[int(k.strip())] = v.strip()
    return ans


def write_template(asks):
    RESOLVE_OUT.parent.mkdir(exist_ok=True)
    L = ["# 확인 필요 항목 — 각 줄 '=' 뒤에 답을 적고 --resolve 로 넘기세요.", ""]
    for t, q in asks:
        L.append(f"# {t['n']:>3}번  {t['date']:%m/%d}  {t['memo']}  {t['amount']:+,}")
        L.append(f"#      {q}")
        L.append(f"{t['n']} = ")
        L.append("")
    RESOLVE_OUT.write_text("\n".join(L), encoding="utf-8")


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        raise Stop(__doc__)
    def opt(flag):
        return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else None

    txns = read_bank(args[0], opt("--pw"))
    start, end = verify_chain(txns)
    print(f"거래 {len(txns)}건 · 잔액 체인 검증 통과")
    print(f"  첫 거래 직전 {start:,}원 → 최종 {end:,}원\n")

    wb, ws, members, retired, months = load_book()
    book = dashboard_balance()
    if start != book:
        raise Stop(
            f"장부와 통장이 어긋납니다.\n"
            f"  통장 첫 거래 직전 : {start:,}원\n"
            f"  현재 장부         : {book:,}원\n"
            f"  차이              : {start-book:+,}원\n"
            f"  → 이 구간에 장부에 없는 거래가 있습니다. 조회기간을 앞으로 늘려 다시 받으세요.\n"
            f"     차액을 임의로 메우지 마십시오.")
    print(f"장부 {book:,}원 = 통장 시작 잔액 · 일치\n")

    fee = standard_fee(ws, members, months)
    answers = read_answers(opt("--resolve"))
    plan, ledger, asks = classify(txns, ws, members, retired, months, fee, answers)

    if asks:
        print(f"■ 확인 필요 {len(asks)}건 (전부 한 번에)")
        for t, q in asks:
            print(f"  {t['n']:>3}번 {t['date']:%m/%d} {t['memo']:16} {t['amount']:>10,}")
            print(f"       └ {q}")
        write_template(asks)
        print(f"\n답안 template 생성 → {RESOLVE_OUT}")
        print("  채운 뒤 --resolve <파일> --apply 로 다시 실행하세요.")
        return

    by = {}
    for _, _, _, name, label, amt in plan:
        by.setdefault(label, []).append(amt)
    print("■ 입금현황 반영 계획")
    for label, vs in by.items():
        print(f"  {label}: {len(vs)}건  +{sum(vs):,}원")
    print(f"  합계 {sum(v for _, _, _, _, _, v in plan):,}원")

    if ledger:
        print("\n■ 원장 추가")
        for d, k1, k2, cha, dae, memo in ledger:
            print(f"  {d}  {k1}/{k2}  {cha or dae:,}  {memo}")

    print(f"\n반영 후 예상 잔액 {end:,}원")
    if "--apply" not in sys.argv:
        print("\n(dry-run) --apply 로 실제 반영")
        return

    for r, c, val, *_ in plan:
        ws.cell(row=r, column=c).value = val
    append_ledger(wb, ledger)
    wb.save(XLSX)
    print(f"\n반영 완료 — 입금현황 {len(plan)}건, 원장 {len(ledger)}행")
    print(f"확인: python finance/src/dashboard.py --check")


if __name__ == "__main__":
    try:
        main()
    except Stop as e:
        print(f"\n중단: {e}")
        sys.exit(1)
