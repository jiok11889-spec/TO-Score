"""
TO-Score 초기 엑셀 파일 생성
"""
import openpyxl
from openpyxl.utils import get_column_letter
import os

MEMBERS = [
    "조현민", "권기표", "옥지엽", "나한영", "류성준", "황인호", "김성환(92)",
    "최원재", "임익준", "김준기", "김산", "김성환", "김지수", "백근호", "손원우",
    "이현우", "이건희", "이도영", "이호준", "황호연", "조성태", "이준범", "서정권",
    "김수민", "이형석", "임승언", "김대우", "이정승", "윤석원", "박헌우", "박병우",
    "홍경택", "윤현호", "강충현", "김서래",
    "윤준원", "김현섭", "강현수", "이은광", "최창환", "주홍석",
]

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'TO-Score.xlsx')


def create_excel():
    wb = openpyxl.Workbook()

    # ── 원장 시트 ──
    ws = wb.active
    ws.title = '원장'

    headers = ['연월', '요일', '일자', '분류1', '분류2', '차변', '대변', '대변-차변', '비고']
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = openpyxl.styles.Font(bold=True)

    ws.cell(2, 3).value = '입력'

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 32

    # ── 입금현황 시트 ──
    ws2 = wb.create_sheet('입금현황')

    ws2.cell(1, 1).value = 'No'
    ws2.cell(1, 1).font = openpyxl.styles.Font(bold=True)
    for col, name in enumerate(MEMBERS, 2):
        c = ws2.cell(1, col, name)
        c.font = openpyxl.styles.Font(bold=True)
    last_col = len(MEMBERS) + 2
    ws2.cell(1, last_col).value = '계'
    ws2.cell(1, last_col).font = openpyxl.styles.Font(bold=True)

    # row 2: 상태 행 (탈퇴 시 해당 셀에 '탈퇴' 입력, 기본 비워둠)

    ws2.column_dimensions['A'].width = 12
    for col in range(2, last_col + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 8

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"[OK] TO-Score.xlsx 생성 완료: {OUTPUT_PATH}")
    print(f"  원장 시트: 헤더 9열 설정")
    print(f"  입금현황 시트: {len(MEMBERS)}명 설정")


if __name__ == '__main__':
    create_excel()
