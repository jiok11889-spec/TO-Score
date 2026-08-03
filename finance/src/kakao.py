"""티오방 카톡 공지 세트 생성 — 이미지 2장 + 문구.

사용법:
    python finance/src/kakao.py

산출물 (finance/out/):
    kakao_1_summary.png   잔액 · 이번달 흐름 · 월별 추이
    kakao_2_payment.png   납부현황 그리드 · 최근 거래
    kakao_message.txt     카톡에 붙여넣을 문구

당월 미납뿐 아니라 **누적 미납**(몇 달째 밀렸는지)을 표시한다.
선납(미래 달까지 미리 낸 사람)은 미납으로 잡지 않는다.
"""
import io
import json
import os
import subprocess
import sys
import time
import urllib.request
from collections import Counter
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

FIN = Path(__file__).resolve().parent.parent
XLSX = FIN / "data" / "TO Score.xlsx"
OUT = FIN / "out"
PORT = 8941
SITE = "https://to-score-finance.onrender.com"


def ym_key(ym):
    y, m = ym.replace("년 ", "-").replace("월", "").strip().split("-")
    return (int(y), int(m))


def read_sheet():
    """입금현황은 행=멤버, 열=월 구조다."""
    ws = openpyxl.load_workbook(XLSX, data_only=True)["입금현황"]
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    months = [(c + 1, str(h).strip()) for c, h in enumerate(header)
              if isinstance(h, str) and "년" in h]
    members = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name or str(name).strip() in ("", "계"):
            continue
        if str(ws.cell(row=r, column=2).value or "").strip() == "탈퇴":
            continue
        pays = {}
        for c, ym in months:
            v = ws.cell(row=r, column=c).value
            pays[ym] = v if isinstance(v, (int, float)) else 0
        members.append({"name": str(name).strip(), "months": pays})
    return [m for _, m in months], members


def build_message(latest, balance, account):
    months, members = read_sheet()
    upto = [m for m in months if ym_key(m) <= ym_key(latest)]

    amts = [m["months"][latest] for m in members if m["months"][latest] > 0]
    fee = Counter(amts).most_common(1)[0][0] if amts else 0

    full, partial, none = [], [], []
    for m in members:
        v = m["months"][latest]
        (full if v >= fee else partial if v > 0 else none).append(
            m["name"] if v >= fee or v == 0 else (m["name"], v))
    partial = [p for p in partial if isinstance(p, tuple)]

    total = sum(m["months"][latest] for m in members)

    # 누적 미납: 첫 납부한 달부터 당월까지, 회비에 못 미친 달
    chronic = []
    for m in members:
        paid_idx = [i for i, ym in enumerate(upto) if m["months"][ym] > 0]
        if not paid_idx:
            continue
        gaps = [ym for ym in upto[paid_idx[0]:] if m["months"][ym] < fee]
        if len(gaps) > 1:
            chronic.append((m["name"], gaps, sum(fee - m["months"][ym] for ym in gaps)))
    chronic.sort(key=lambda x: -len(x[1]))

    mm = latest.split()[-1]
    L = [f"티오방 {mm} 회비입출내역입니다 ⛳", "",
         f"▪ 현재 잔액 : {balance:,}원",
         f"▪ {mm} 입금 : {total:,}원 ({len(full) + len(partial)}명 / {len(members)}명)", ""]
    if none:
        L.append(f"▪ {mm} 미납")
        for i in range(0, len(none), 4):
            L.append(" ".join(none[i:i + 4]))
        L.append("")
    for n, v in partial:
        L.append(f"※ {n}님 {v:,}원 부분납부 (잔여 {fee - v:,}원)")
    for n, gaps, short in chronic:
        L.append(f"※ {n}님 {len(gaps)}개월 미납 ({', '.join(g.split()[-1] for g in gaps)}) — 총 {short:,}원")
    if partial or chronic:
        L.append("")
    L += ["▪ 입금계좌",
          f"카카오뱅크 {account['number']} ({account['holder']})", "", SITE]
    return "\n".join(L)


BOXES = """() => {
  const secs = [...document.querySelectorAll('.sec')];
  const txt = e => (e.innerText || '').trim().toUpperCase();
  const flow = secs.find(s => txt(s).includes('MONTHLY FLOW'));
  const hero = document.querySelector('.hero');
  const two  = document.querySelector('.two');
  const r = e => { const b = e.getBoundingClientRect();
                   return {x: b.left + scrollX, y: b.top + scrollY, w: b.width, h: b.height}; };
  return {flow: r(flow), hero: r(hero), two: r(two), docW: document.body.scrollWidth};
}"""
PAD = 28


def capture(url):
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        b = p.chromium.launch()
        pg = b.new_page(viewport={"width": 900, "height": 1400}, device_scale_factor=2)
        pg.goto(url, wait_until="networkidle")
        pg.wait_for_timeout(3000)
        bx = pg.evaluate(BOXES)
        x = max(0, bx["hero"]["x"] - PAD)
        w = min(bx["docW"] - x, bx["hero"]["w"] + PAD * 2)

        def shot(name, top, bottom):
            pg.screenshot(path=str(OUT / name), full_page=True,
                          clip={"x": x, "y": top, "width": w, "height": bottom - top})

        shot("kakao_1_summary.png", max(0, bx["hero"]["y"] - PAD),
             bx["flow"]["y"] + bx["flow"]["h"] + PAD)
        shot("kakao_2_payment.png", bx["two"]["y"] - PAD,
             bx["two"]["y"] + bx["two"]["h"] + PAD)
        b.close()


def main():
    OUT.mkdir(exist_ok=True)
    env = dict(os.environ, PORT=str(PORT))
    server = subprocess.Popen([sys.executable, str(FIN / "src" / "dashboard.py")],
                              env=env, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    url = f"http://localhost:{PORT}"
    try:
        for _ in range(40):
            try:
                with urllib.request.urlopen(url + "/api/data") as r:
                    d = json.load(r)
                break
            except Exception:
                time.sleep(0.5)
        else:
            raise SystemExit("대시보드 서버가 뜨지 않았습니다.")

        msg = build_message(d["summary"]["latest_ym"], round(d["summary"]["balance"]),
                            d["account"])
        (OUT / "kakao_message.txt").write_text(msg, encoding="utf-8")
        capture(url)
    finally:
        server.terminate()

    print(msg)
    print(f"\n→ {OUT}")


if __name__ == "__main__":
    main()
